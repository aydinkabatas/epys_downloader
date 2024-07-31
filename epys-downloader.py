from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import math
import pandas as pd
import os, fnmatch

# Login Info
usernm = "USERNAME"
passwd = "PASSWORD"

# Find Desktop Path And Meter ID File
desktop_path = os.path.expanduser("~/Desktop")
main_file_path = desktop_path+'/scriptie4.xlsx'
daf = pd.read_excel(main_file_path)
daf["SAYACID"] = daf["SAYACID"].astype('str')

def khan(): # EPUI-Loader holder.
    while True:
        try:
            b2c=drv.find_element(By.CLASS_NAME,"epui-loader")
            time.sleep(1)
        except:        
            break;

def finder(pattern, path): # File Finder
    for root, folder, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result=os.path.join(root, name)
                break
    return result

# Login Page
drv = webdriver.Chrome()
url="https://cas.epias.com.tr/cas/login"
drv.get(url)

securedoor = drv.find_element(By.ID, "username")
securedoor.clear()
securedoor.send_keys(usernm)
securedoor = drv.find_element(By.ID, "password")
securedoor.clear()
securedoor.send_keys(passwd)

b1c = drv.find_element(By.ID, "login-btn2") # Login Button Press
b1c.click()
time.sleep(0.5)

# EPIAS CAS Entrance Key (Hold Until 5 Digit Entrance) - DONT PRESS ENTER
while True:
    b1c = drv.find_element(By.ID, "token").get_attribute("value")
    if len(b1c)==5:
        break
    time.sleep(0.5)

b1c = drv.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div/form/div[3]/button[1]") # Authenticate Button Press
b1c.click()
time.sleep(0.5)


# Redirected to: Reconciliation Operations > Data Operation > Approved Meter Data
url="https://epys.epias.com.tr/reconciliation-operations/data-operation/approved-meter-data"
drv.get(url)
khan()

# Set Variables
meter_series = []
meter_series = daf["SAYACID"].tolist()
count_m = len(meter_series)
part = 0
counter = 0
max_part = math.ceil(count_m/10)

# Start Loop For Meter Download
for i in range(1, max_part+1): # EPYS System Only Take 10 Meter ID Every Query. We Have Divided All Meter ID's Into Parts.

    if max_part==i:
        add_loop=count_m%10
    else:
        add_loop=10

    try: # If Error, Skip These Meter ID's. You Can Catch Skipped Part On ControlFile Page. 
        for j in range(part,part+add_loop): # Loop In Part Inside. This Loop Make Every Meter ID Entrance One By One.
            b2c=drv.find_element(By.ID, "meterId")
            b2c.send_keys(meter_series[j],Keys.ENTER)
            counter+=1
        b2c=drv.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/div/div[2]/div[12]/button") # Submit Meter ID's
        b2c.click()

        khan()

        for j in range(add_loop): # Loop For Downloading Every Meter ID.
            try: # If Error, Skip This Meter ID's. You Can Catch Skipped Part On ControlFile Page. 
                meter_link="/html/body/div[1]/div/main/div[1]/main/section/div[3]/div/div/div[2]/div[2]/div/div/div["+str(j+1)+"]/div/div/div[6]/button" # Select Meter ID
                b2c=drv.find_element(By.XPATH, meter_link)
                b2c.click()
                time.sleep(0.5)
                b2c=drv.find_element(By.XPATH, '/html/body/div[4]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div[1]') # Select Options Button
                b2c.click()
                b2c=drv.find_element(By.XPATH, '/html/body/div[4]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[1]') # Select "Download via Excel File"
                b2c.click()
                time.sleep(0.5)
                b2c=drv.find_element(By.XPATH, '/html/body/div[4]/div/div/div[1]/button') # Close Subpage
                b2c.click()
            except:
                break

        b2c=drv.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/div/div[2]/div[10]/button") # Clear Old Meter ID's
        b2c.click()
        part=counter

    except:
        b2c=drv.find_element(By.XPATH, '/html/body/div[4]/div/div/div[1]/button') # Close Error Subpage
        b2c.click()
        time.sleep(1)
        b2c=drv.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/div/div[2]/div[10]/button") # Clear Old Meter ID's
        b2c.click()
        time.sleep(1)
        part=counter

# Find Download File
download_path = os.path.expanduser("~/Downloads/")

# Control All Downloaded Hourly Meters
for i in range(daf["SAYACID"].size):
    try:
        pattern_search=str(daf["SAYACID"][i])+"_Saatlik*"
        meterfilename=finder(pattern_search, download_path)
    except:
        meterfilename='File Not Exists.'
    
    daf.at[i, "File Path"] = meterfilename #Create A New Header


# Create ControlFile Page
with pd.ExcelWriter(main_file_path, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
    daf.to_excel(writer, sheet_name='ControlFile', index=False)

# Create HourlyPivot Page

wb = load_workbook(main_file_path)
if "HourlyPivot" in wb.sheetnames:
        del wb["HourlyPivot"]
new_sheet = wb.create_sheet(title="HourlyPivot")

column = 2
# Loop For Every Meter ID
for filename in daf["File Path"]:
    if filename != 'File Not Exists.': # Control For File Exists.
        hourly_wb = load_workbook(filename)
        hourly_sheet = hourly_wb.active
        cell_value = hourly_sheet[f'C2'].value # Get Header For Every Counter ID
        new_sheet.cell(row=1, column=column, value=cell_value)
        for row in range(2, hourly_sheet.max_row+1): # Get Every Hourly Consumption. Side by side. 
            cell_value = float(hourly_sheet[f'G{row}'].value) * 1000 # Convert to MwH > KwH 
            new_sheet.cell(row=row, column=column, value=cell_value)
        hourly_wb.close()
        column += 1  # Next Counter ID

wb.save(main_file_path) # Save Same File

print("-------------------------------------------")
print("Congratulations! Process complete without errors!")
